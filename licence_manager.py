from tkinter import Tk, Frame, Label, Entry, Button, StringVar, ttk
import os
import requests
import json
from dotenv import load_dotenv
from datetime import datetime

# Import for Excel file management
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("Veuillez installer openpyxl (pip install openpyxl) pour activer la journalisation Excel.")

load_dotenv()  # Charge le contenu du fichier .env

API_URL = "https://license-api-h5um.onrender.com/modify_credits"
SECRET = os.getenv("ADMIN_PASSWORD")  # ADMIN_PASSWORD is your secret token

# ---------------------------- Excel Logger Class ----------------------------
class ExcelLogger:
    def __init__(self, file="update_log.xlsx"):
        self.file = file
        self.header = ["Date/Heure", "Client", "Balance Type", "Item Name", "Amount"]
        if not os.path.exists(self.file):
            wb = Workbook()
            ws = wb.active
            ws.append(self.header)
            wb.save(self.file)

    def log(self, client, balance_type, item_name, amount):
        try:
            wb = load_workbook(self.file)
            ws = wb.active
            ws.append([
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                client,
                balance_type,
                item_name,
                amount
            ])
            wb.save(self.file)
        except Exception as e:
            print("Erreur lors de la journalisation :", e)

# Global ExcelLogger instance shared by the application
excel_logger = ExcelLogger()

# -------------------------- Tkinter License Manager -------------------------
class LicenceManagerFrame(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.pack(padx=10, pady=10)

        self.agency_var = StringVar()
        self.amount_matrix_var = StringVar()
        self.amount_weighter_var = StringVar()
        self.new_agency_var = StringVar()
        self.item_name_var = StringVar()  # For additional flexibility if needed

        # 🔴 Ligne 0 — Label d'avertissement
        self.result_label = Label(self, text="LOTS EN MILLIERS DE MOTS!!!!", fg="red")
        self.result_label.grid(row=0, column=0, columnspan=5, sticky="w", pady=(0, 10))

        # 🧾 Ligne 1 — Choix de l'agence
        Label(self, text="Agence:").grid(row=1, column=0, sticky="w")
        self.agency_combobox = ttk.Combobox(self, textvariable=self.agency_var, values=[], state="readonly")
        self.agency_combobox.bind("<<ComboboxSelected>>", self.check_current_balance)
        self.agency_combobox.grid(row=1, column=1, columnspan=2, sticky="ew")

        # 📦 Ligne 2 — Solde Matriciel
        Label(self, text="Solde Matriciel").grid(row=2, column=0, sticky="w")
        self.matrix_balance_label = Label(self, text="0 crédits")
        self.matrix_balance_label.grid(row=2, column=1, sticky="w")
        self.matrix_entry = Entry(self, textvariable=self.amount_matrix_var, width=10)
        self.matrix_entry.grid(row=2, column=2, sticky="w")
        Button(self, text="Ajouter", command=self.add_matrix_credits).grid(row=2, column=3)
        Button(self, text="Retirer", command=self.remove_matrix_credits).grid(row=2, column=4)

        # ⚖️ Ligne 3 — Solde Pondérateur
        Label(self, text="Solde Pondérateur").grid(row=3, column=0, sticky="w")
        self.weighter_balance_label = Label(self, text="0 crédits")
        self.weighter_balance_label.grid(row=3, column=1, sticky="w")
        self.weighter_entry = Entry(self, textvariable=self.amount_weighter_var, width=10)
        self.weighter_entry.grid(row=3, column=2, sticky="w")
        Button(self, text="Ajouter", command=self.add_weighter_credits).grid(row=3, column=3)
        Button(self, text="Retirer", command=self.remove_weighter_credits).grid(row=3, column=4)

        # 🆕 Ligne 4 — Nouvelle agence
        Label(self, text="Nouvelle agence").grid(row=4, column=0, sticky="w", pady=(10, 0))
        self.new_agency_entry = Entry(self, textvariable=self.new_agency_var)
        self.new_agency_entry.grid(row=4, column=1, columnspan=2, sticky="ew", pady=(10, 0))
        Button(self, text="Ajouter agence", command=self.create_agency).grid(row=4, column=3, columnspan=2, pady=(10, 0))

        # Permet l'expansion horizontale
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)

        self.fetch_agencies()

    def update_credits(self, amount, balance_type, item_name=""):
        """
        Mise à jour des crédits pour l'agence courante.
        Les informations envoyées sont : agency_name, amount, balance_type,
        et item_name. Un enregistrement de cet appel est ajouté dans le fichier Excel.
        Lorsque le token est SECRET, le client est "Administration".
        """
        # Log the update with client = "Administration"
        excel_logger.log("Administration", balance_type, item_name, amount)

        try:
            headers = {
                "Authorization": f"Bearer {SECRET}"
            }

            data = {
                "agency_name": self.agency_var.get(),
                "amount": amount,
                "balance_type": f"{balance_type}_balance",
                "item_name": item_name
            }

            print("📤 Requête envoyée :")
            print(json.dumps(data, indent=2, ensure_ascii=False))
            print("🔐 Headers :", headers)

            response = requests.post(API_URL, json=data, headers=headers)

            print("📥 Réponse brute :")
            print("Status code :", response.status_code)
            print("Texte brut :", response.text)

            try:
                result = response.json()
            except ValueError:
                self.result_label.config(text=f"❌ Réponse invalide : {response.text}")
                return

            if result.get("success"):
                self.result_label.config(text="✅ Crédits mis à jour.")
                self.check_current_balance()  # Mise à jour des soldes affichés
            else:
                self.result_label.config(text=f"❌ Erreur: {result.get('error')}")
        except Exception as e:
            self.result_label.config(text=f"❌ Erreur réseau: {e}")

    def check_current_balance(self, *_):
        try:
            agency_name = self.agency_var.get()
            headers = {
                "Authorization": f"Bearer {agency_name}"
            }
            response = requests.post("https://license-api-h5um.onrender.com/get_balance", headers=headers)

            if response.status_code != 200:
                self.result_label.config(text=f"❌ Erreur HTTP {response.status_code} : {response.text}")
                return

            if not response.content:
                self.result_label.config(text="❌ Réponse vide du serveur.")
                return

            result = response.json()

            if result.get("success"):
                matrix = result.get("matrix_balance", "?")
                weighter = result.get("weighter_balance", "?")

                self.matrix_balance_label.config(text=f"{matrix} crédits")
                self.weighter_balance_label.config(text=f"{weighter} crédits")
                self.result_label.config(text="")  # Efface les anciens messages d’erreur
            else:
                self.result_label.config(text=f"❌ Erreur: {result.get('error')}")
        except Exception as e:
            self.result_label.config(text=f"❌ Erreur réseau: {e}")

    def create_agency(self):
        agency_name = self.new_agency_var.get().strip()
        if not agency_name:
            self.result_label.config(text="⚠️ Entrez un nom d’agence.")
            return

        headers = {
            "Authorization": f"Bearer {SECRET}"
        }
        data = {
            "agency_name": agency_name,
            "matrix_balance": 0,
            "weighter_balance": 0
        }

        try:
            response = requests.post("https://license-api-h5um.onrender.com/add_agency", json=data, headers=headers)
            if response.content:
                result = response.json()
            else:
                self.result_label.config(text="❌ Erreur : réponse vide du serveur.")
                return
            if result.get("success"):
                self.result_label.config(text=f"✅ Agence '{agency_name}' ajoutée.")
                self.fetch_agencies()
                self.agency_var.set(agency_name)
            else:
                self.result_label.config(text=f"❌ Erreur: {result.get('error')}")
        except Exception as e:
            self.result_label.config(text=f"❌ Erreur réseau: {e}")

    def fetch_agencies(self):
        try:
            response = requests.get("https://license-api-h5um.onrender.com/list_agencies")
            response.raise_for_status()
            result = response.json()

            if isinstance(result, list):
                self.agencies = result
            elif isinstance(result, dict):
                self.agencies = result.get("agencies", [])
            else:
                raise ValueError("Format inattendu de la réponse")

            self.agency_combobox['values'] = self.agencies

            if self.agencies:
                self.agency_var.set(self.agencies[0])
        except Exception as e:
            self.result_label.config(text=f"❌ Impossible de récupérer les agences: {e}")

    def add_matrix_credits(self):
        try:
            amount = int(self.amount_matrix_var.get()) * 1000
            # "matrix" is used to match matrix_balance variable name.
            self.update_credits(amount=amount, balance_type='matrix', item_name="Crédit")
        except ValueError:
            self.result_label.config(text="⚠️ Entrez un nombre valide pour le solde matriciel.")

    def remove_matrix_credits(self):
        try:
            amount = int(self.amount_matrix_var.get()) * -1000
            # "matrix" is used to match matrix_balance variable name.
            self.update_credits(amount=amount, balance_type='matrix', item_name="Débit")
        except ValueError:
            self.result_label.config(text="⚠️ Entrez un nombre valide pour le solde matriciel.")

    def add_weighter_credits(self):
        try:
            amount = int(self.amount_weighter_var.get()) * 1000
            # "weigher" is used to match weighter_balance variable name.
            self.update_credits(amount=amount, balance_type='weighter', item_name="Crédit")
        except ValueError:
            self.result_label.config(text="⚠️ Entrez un nombre valide pour le solde pondérateur.")

    def remove_weighter_credits(self):
        try:
            amount = int(self.amount_weighter_var.get()) * -1000
            # "weigher" is used to match weighter_balance variable name.
            self.update_credits(amount=amount, balance_type='weighter', item_name="Débit")
        except ValueError:
            self.result_label.config(text="⚠️ Entrez un nombre valide pour le solde pondérateur.")


if __name__ == "__main__":
    root = Tk()
    root.title("Gestion des crédits")
    frame = LicenceManagerFrame(master=root)
    root.mainloop()
